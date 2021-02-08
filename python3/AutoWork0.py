#!/usr/bin/env python3
# -*- coding: utf-8 -*-
import codecs
import contextlib
import datetime
import os
import sys
from typing import List, Tuple
import sqlite3 as s3
from exchangelib import DELEGATE, Credentials, Account
from bs4 import BeautifulSoup
from openpyxl import Workbook, load_workbook

from openpyxl.worksheet import worksheet
from sqlite3 import Cursor


class Seq_Struct:

    def __init__(self):
        self.__seq = ['bad_lib_s', 'bad_proto_s', 'stora_wgit_s', 'more_tag_s', 'com_acc_s']
        for name in self.__seq:
            setattr(self, name, list())

    def get_sorted_seq(self):
        return self.__seq

    def to_list(self):
        return [getattr(self, name) for name in self.__seq]


__today_date_str__ = datetime.datetime.strftime(datetime.datetime.now(), '%Y/%m/%d')
__today_date_str__1 = datetime.datetime.strftime(datetime.datetime.now(), '%Y-%m-%d')


def is_over_dur_(curr_date: str, prev_date: str, delta: int = 14, formater: str = '%Y/%m/%d') -> bool:
    c = datetime.datetime.strptime(curr_date, formater)
    p = datetime.datetime.strptime(prev_date, formater)
    return (c - p).days >= 14


@contextlib.contextmanager
def get_cursor(db: str) -> Cursor:
    conn = None
    try:
        conn = s3.connect(db)
        cur = conn.cursor()
        yield cur
        conn.commit()
    except Exception as e:
        print(str(e))
        conn.rollback()
        raise cur
    finally:
        if conn is not None:
            conn.close()


def select_(db_file: str, sql: str, params: Tuple) -> List:
    ps = list()
    for p in params:
        if isinstance(p, str):
            ps.append(_stand_strip_br(p))
        else:
            ps.append(p)
    with get_cursor(db_file) as cur:
        cur.execute(sql, tuple(ps))
        return cur.fetchall()
    return None


def execute_(db_file: str, sql: str, params: Tuple):
    ps = list()
    for p in params:
        if isinstance(p, str):
            ps.append(_stand_strip_br(p))
        else:
            ps.append(p)
    with get_cursor(db_file) as cur:
        cur.execute(sql, tuple(ps))


def save_many_(db_file: str, sql: str, params_s: List[Tuple]):
    with get_cursor(db_file) as cur:
        cur.executemany(sql, params_s)


def strip(s: str) -> str:
    return '' if s is None else str(s).strip()


def _stand_strip(s: str) -> str:
    s = strip(s)
    return s.replace('\r\n', chr(10))


def _stand_strip_br(s: str) -> str:
    s = strip(s)
    return s.replace('\n', '<tadus_br>')


def _imstand_strip_br(s: str) -> str:
    s = strip(s)
    return s.replace('<tadus_br>', '\n')


def _stand_strip_w(s: str) -> str:
    s = _stand_strip(s)
    return s.replace('\n', '\r\n')


credentials = Credentials(r'tencent.com\v_fantnzeng', 'zs9709ZSB')

account = Account('v_fantnzeng@tencent.com', credentials=credentials, autodiscover=True)


def mtx_to_sht(sht: worksheet, mtx: List[List]):
    for row in range(len(mtx)):
        for col in range(len(mtx[row])):
            value = strip(mtx[row][col])
            sht.cell(row=row + 1, column=col + 1, value=value)


def mtx_to_wb(wb: Workbook, sht_name: str, idx: int, mtx: List[List]):
    wb.create_sheet(sht_name, idx)
    sht = wb[sht_name]
    mtx_to_sht(sht, mtx)


def save_all_(file_path: str, all_: Seq_Struct):
    wb = Workbook()
    s = all_.get_sorted_seq()
    for i in range(len(s)):
        mtx_to_wb(wb, s[i], i, getattr(all_, s[i]))
    wb.save(file_path)


def get_mail_by_subject(subject_name: str) -> str:
    mails = account.inbox.filter(subject__contains=subject_name)
    for mail in mails:
        if str(mail.subject).strip() == str(subject_name).strip():
            return mail.body
    return None


def simple_find_table_0(soup, title_name: str):
    bs = soup.select('font')
    for b in bs:
        if b.get_text().strip() == title_name:
            table = b.find_next_sibling().find_next_sibling()
            return table
    return None


def simple_find_table_1(soup, title_name: str):
    bs = soup.select('font')
    for b in bs:
        if b.get_text().strip() == title_name:
            table = b.find_next_sibling()
            return table
    return None


# 【微信支付-日报】【理想家园】 编译加速系统运营日报
def get_ComAcc_mail(date_str_s: List[str]) -> List[str]:
    title_ = '【微信支付-日报】【理想家园】 编译加速系统运营日报'
    return [get_mail_by_subject(f'{title_} {date}') for date in date_str_s]


def souptable_to_matrix(table) -> List[List[str]]:
    res = list()
    trs = table.select('tbody > tr')
    for tr in trs:
        ths = tr.select('th')
        if len(ths) > 0:
            th_s = [_stand_strip(th.get_text()) for th in ths]
            res.append(th_s)
            continue
        tds = tr.select('td')
        if len(tds) > 0:
            td_s = [_stand_strip(td.get_text()) for td in tds]
            res.append(td_s)
            continue
    return res


def get_ComAcc_htmltable(date_str_s: List[str]) -> List:
    mails = get_ComAcc_mail(date_str_s)
    soups = [BeautifulSoup(mail, 'html.parser') for mail in mails]
    return [simple_find_table_0(soup, '用户patchbuild编译次数Top榜') for soup in soups]


def get_ComAcc_matrix(date_str_s: List[str]) -> List[List[str]]:
    souptables = get_ComAcc_htmltable(date_str_s)
    return [souptable_to_matrix(souptable) for souptable in souptables]


# 【微信支付】【理想家园】 项目全局简报
def get_ProjectGlobal_mail(date_str_s: List[str]) -> List[str]:
    title_ = '【微信支付】【理想家园】 项目全局简报'
    return [get_mail_by_subject(f'{title_}[{date}]') for date in date_str_s]


def get_ProjectGlobal_soup(date_str_s: List[str]) -> List:
    mails = get_ProjectGlobal_mail(date_str_s)
    return [BeautifulSoup(mail, 'html.parser') for mail in mails]


def get_matrix_by_soup_0(soups: List, title_name) -> List:
    tables = [simple_find_table_0(soup, title_name) for soup in soups]
    return [souptable_to_matrix(table) for table in tables]


def get_matrix_by_soup_1(soups: List, title_name) -> List:
    tables = [simple_find_table_1(soup, title_name) for soup in soups]
    return [souptable_to_matrix(table) for table in tables]


def get_BadLibMatrix_by_soup(ProjectGlobal_soups: List) -> List:
    return get_matrix_by_soup_0(ProjectGlobal_soups, '● 昨日新增违规文件')


def get_StoraWGitMatrix_by_soup(ProjectGlobal_soups: List) -> List:
    return get_matrix_by_soup_1(ProjectGlobal_soups, '● 过去10天新增模块接入整洁GIT情况')


def get_TooMoreTagMatrix_by_soup(ProjectGlobal_soups: List) -> List:
    return get_matrix_by_soup_0(ProjectGlobal_soups, '微信支付部分接入整洁GIT编译依赖数量Top30')


def get_BadProtoMatrix_by_soup(ProjectGlobal_soups: List) -> List:
    return get_matrix_by_soup_0(ProjectGlobal_soups, '● Proto文件规范化检查')


# date_str like %Y-%m-%d
def get_global_mtx(date_str: List[str]) -> Seq_Struct:
    date_s = [strip(date).split('-') for date in date_str]
    date_s = [[int(i) for i in j] for j in date_s]
    date_s = ['%d-%02d-%02d' % (date[0], date[1], date[2]) for date in date_s]
    res = Seq_Struct()
    soup_s = get_ProjectGlobal_soup(date_s)
    res.bad_lib_s = get_BadLibMatrix_by_soup(soup_s)
    res.stora_wgit_s = get_StoraWGitMatrix_by_soup(soup_s)
    res.more_tag_s = get_TooMoreTagMatrix_by_soup(soup_s)
    res.bad_proto_s = get_BadProtoMatrix_by_soup(soup_s)
    return res


# date_str like %Y-%m-%d
def get_com_acc_mtx(date_str: List[str]) -> List[List]:
    date_s = [strip(date).split('-') for date in date_str]
    date_s = [[int(i) for i in j] for j in date_s]
    date_s = ['%d%02d%02d' % (date[0], date[1], date[2]) for date in date_s]
    return get_ComAcc_matrix(date_s)


# date_str like %Y-%m-%d
def do_get_all(date_str: List[str]) -> Seq_Struct:
    def _in_func_0(mtx: List[List]) -> List[List[List]]:
        for i in range(1, len(mtx)):
            mtx[i][1] = ';'.join(mtx[i][1].split(','))
        return [mtx]

    res = get_global_mtx(date_str)
    res.bad_lib_s = _in_func_0(res.bad_lib_s[0])
    res.com_acc_s = get_com_acc_mtx(date_str)
    return res


tail_hand = ['是否通知', '备注', '上次通知时间', '是否需要通知']
default_tail = ['否', '', __today_date_str__, '是']


def g_handle_badlib(db_file: str, mtx: List[List]) -> List[List]:
    sql = 'select id,is_need_noti,prev_noti_date from t_bad_lib where file=? and lines=? and not_xlib=? and tobe_xlib=? and those=?'
    hand = mtx[0][:] + tail_hand
    res = [hand]
    for i in range(1, len(mtx)):
        a = select_(db_file, sql, tuple(mtx[i]))
        if len(a) <= 0 or \
                (a[0][1] != 0 and is_over_dur_(default_tail[2], a[0][2])):
            res.append(mtx[i] + default_tail)
            continue
    return res


def g_handle_badproto(db_file: str, mtx: List[List]) -> List[List]:
    sql = 'select id,is_need_noti,prev_noti_date from t_bad_proto where those=? and problem_proto=? and not_rule=?'
    hand = mtx[0][:] + tail_hand
    res = [hand]
    for i in range(1, len(mtx)):
        a = select_(db_file, sql, tuple(mtx[i][1:]))
        if len(a) <= 0 or \
                (a[0][1] != 0 and is_over_dur_(default_tail[2], a[0][2])):
            res.append(mtx[i] + default_tail)
            continue
    return res


def g_handle_storawgit(db_file: str, mtx: List[List]) -> List[List]:
    sql = 'select id,is_need_noti,prev_noti_date from t_stora_wgit where module_name=? and module_those=?'
    hand = mtx[0][:] + tail_hand
    res = [hand]
    for i in range(1, len(mtx)):
        if strip(mtx[i][1]) != '未入整洁GIT':
            continue
        a = select_(db_file, sql, (mtx[i][0], mtx[i][3]))
        if len(a) <= 0 or \
                (a[0][1] != 0 and is_over_dur_(default_tail[2], a[0][2])):
            res.append(mtx[i] + default_tail)
            continue
    return res


def g_handle_comacc(db_file: str, mtx: List[List]) -> List[List]:
    sql = 'select id,is_need_noti,prev_noti_date from t_com_acc where account=?'
    hand = mtx[0][:] + tail_hand
    res = [hand]
    for i in range(1, len(mtx)):
        if strip(strip(mtx[i][3])).upper() != 'NO':
            continue
        a = select_(db_file, sql, (mtx[i][0],))
        if len(a) <= 0 or \
                (a[0][1] != 0 and is_over_dur_(default_tail[2], a[0][2])):
            res.append(mtx[i] + default_tail)
            continue
    return res


def g_handle_toomorede(db_file: str, mtx: List[List]) -> List[List]:
    sql = 'select id,is_need_noti,prev_noti_date from t_too_more_target where file=? and target_name=? and those=?'
    hand = mtx[0][:] + tail_hand
    res = [hand]
    for i in range(1, len(mtx)):
        a = select_(db_file, sql, (mtx[i][0], mtx[i][1], mtx[i][3]))
        if len(a) <= 0 or \
                (a[0][1] != 0 and is_over_dur_(default_tail[2], a[0][2])):
            res.append(mtx[i] + default_tail)
            continue
    return res


def g_handle_all(db_file: str, all_: Seq_Struct) -> Seq_Struct:
    all_.com_acc_s = g_handle_comacc(db_file, all_.com_acc_s[0])
    all_.bad_lib_s = g_handle_badlib(db_file, all_.bad_lib_s[0])
    all_.bad_proto_s = g_handle_badproto(db_file, all_.bad_proto_s[0])
    all_.stora_wgit_s = g_handle_storawgit(db_file, all_.stora_wgit_s[0])
    all_.more_tag_s = g_handle_toomorede(db_file, all_.more_tag_s[0])
    return all_


def gen_html_(file_path: str, hand: List, context: List[List]):
    if len(context) <= 0:
        return
    h_s = '\n'.join([f'<th>{h}</th>' for h in hand])
    tr = '\n'.join([f'<tr>{td}</tr>' for td in
                    ['\n'.join([f'<td>{row[c_ix]}</td>' for c_ix in range(len(hand))]) for row in context]])
    table = f'<table border="1" style="border-collapse: collapse;">\n<tr style="background-color:#eff7ff;">{h_s}</tr>\n{tr}'
    template = f'''
<html>
<head>
    <meta http-equiv="Content-Type" content="text/html; charset=utf-8">
</head>
<body>
{table}
</body>
</html>
        '''
    with codecs.open(file_path, "w", "utf-8") as f:
        f.write(template)


def gen_badlib_report(mtx: List[List], out_dir: str):
    out_dir_ = f'{out_dir}{os.path.sep}badlib_s{os.path.sep}'
    if not os.path.exists(out_dir_):
        os.mkdir(out_dir_)
    handle = mtx[0][:-4]
    cache = dict()
    for i in range(1, len(mtx)):
        name = mtx[i][4]
        if name not in cache:
            cache[name] = list()
        cache[name].append(mtx[i])
    for k, v in cache.items():
        gen_html_(f'{out_dir_}XLIB违规文件-{k}-{__today_date_str__1}.html', handle, v)


def gen_badproto_report(mtx: List[List], out_dir: str):
    out_dir_ = f'{out_dir}{os.path.sep}badproto_s{os.path.sep}'
    if not os.path.exists(out_dir_):
        os.mkdir(out_dir_)
    handle = mtx[0][:-4]
    cache = dict()
    for i in range(1, len(mtx)):
        name = mtx[i][1]
        if name not in cache:
            cache[name] = list()
        cache[name].append(mtx[i])
    for k, v in cache.items():
        gen_html_(f'{out_dir_}Proto文件规范化检查-{k}-{__today_date_str__1}.html', handle, v)


def gen_storawgit_report(mtx: List[List], out_dir: str):
    out_dir_ = f'{out_dir}{os.path.sep}storawgit_s{os.path.sep}'
    if not os.path.exists(out_dir_):
        os.mkdir(out_dir_)
    handle = mtx[0][:-4]
    cache = dict()
    for i in range(1, len(mtx)):
        name = mtx[i][3]
        if name not in cache:
            cache[name] = list()
        cache[name].append(mtx[i])
    for k, v in cache.items():
        gen_html_(f'{out_dir_}新增模块接入整洁GIT情况-{k}-{__today_date_str__1}.html', handle, v)


def gen_toomorede_report(mtx: List[List], out_dir: str):
    out_dir_ = f'{out_dir}{os.path.sep}toomorede_s{os.path.sep}'
    if not os.path.exists(out_dir_):
        os.mkdir(out_dir_)
    handle = mtx[0][:-4]
    cache = dict()
    for i in range(1, len(mtx)):
        name = mtx[i][3]
        if name not in cache:
            cache[name] = list()
        cache[name].append(mtx[i])
    for k, v in cache.items():
        gen_html_(f'{out_dir_}微信支付部分接入整洁GIT编译依赖数量-{k}-{__today_date_str__1}.html', handle, v)


def gen_comacc_report(mtx: List[List], out_dir: str):
    out_dir_ = f'{out_dir}{os.path.sep}comacc_s{os.path.sep}'
    if not os.path.exists(out_dir_):
        os.mkdir(out_dir_)
    handle = mtx[0][:-4]
    cache = dict()
    for i in range(1, len(mtx)):
        name = mtx[i][0]
        if name not in cache:
            cache[name] = list()
        cache[name].append(mtx[i])
    for k, v in cache.items():
        gen_html_(f'{out_dir_}用户patchbuild编译-{k}-{__today_date_str__1}.html', handle, v)


def gen_all_reports(all_: Seq_Struct, out_dir: str):
    if os.path.exists(out_dir):
        if os.path.isdir(out_dir):
            pass
        else:
            raise out_dir
    else:
        os.mkdir(out_dir)
    # ['bad_lib_s', 'bad_proto_s', 'stora_wgit_s', 'more_tag_s', 'com_acc_s']
    gen_badlib_report(all_.bad_lib_s, out_dir)
    gen_badproto_report(all_.bad_proto_s, out_dir)
    gen_toomorede_report(all_.more_tag_s, out_dir)
    gen_storawgit_report(all_.stora_wgit_s, out_dir)
    gen_comacc_report(all_.com_acc_s, out_dir)


def genreport(db_file: str = None, out_dir: str = None, in_date: str = None):
    if in_date is None:
        in_date = datetime.datetime.strftime(datetime.datetime.now(), '%Y-%m-%d')
    if db_file is None:
        db_file = f'{sys.path[0]}{os.path.sep}IdealHome.db3'
    if out_dir is None:
        out_dir = f'{in_date}-out'
    print(in_date, db_file, out_dir)
    all_ = do_get_all([in_date])
    all_ = g_handle_all(db_file, all_)
    gen_all_reports(all_, out_dir)
    out_xlsx = f'{out_dir}{os.path.sep}__out.xlsx'
    save_all_(out_xlsx, all_)


def get_sht_head(sht: worksheet) -> List[str]:
    head = list()
    col_num = 0
    while True:
        col_num += 1
        c = strip(sht.cell(row=1, column=col_num).value)
        if c == '':
            break
        else:
            head.append(c)
    return head


def read_from_sht(sht: worksheet) -> List[List[str]]:
    head = get_sht_head(sht)
    res = [head]
    begin_p = 1
    while True:
        begin_p += 1
        row = list()
        c_ = 0
        for c in range(len(head)):
            v = strip(sht.cell(row=begin_p, column=c + 1).value)
            row.append(v)
            if v != '':
                c_ += 1
        if c_ <= 0:
            break
        res.append(row)
    return res


def read_db_xlsx(file_path: str) -> Seq_Struct:
    wb = load_workbook(file_path)
    all_ = Seq_Struct()
    s = all_.get_sorted_seq()
    for sheet_name in s:
        sht = wb[sheet_name]
        setattr(all_, sheet_name, read_from_sht(sht))
    return all_


def update_badlib(db_file: str, mtx: List[List[str]]):
    find_sql = 'select id from t_bad_lib where file=? and lines=? and not_xlib=? and tobe_xlib=? and those=?'
    update_sql = 'update t_bad_lib set is_noti=?,noti=?,prev_noti_date=?,is_need_noti=? where id=?'
    insert_sql = r'insert into t_bad_lib ' \
                 r'(file,lines,not_xlib,tobe_xlib,those,is_noti,noti,prev_noti_date,is_need_noti)' \
                 r' values(?,?,?,?,?,?,?,?,?);'
    for i in range(1, len(mtx)):
        mtx[i][-4] = 1 if mtx[i][-4] == '是' else 0  # is_noti
        mtx[i][-1] = 0 if mtx[i][-1] == '否' else 1  # is_need_noti
        a = select_(db_file, find_sql, tuple(mtx[i][:-4]))
        if len(a) <= 0:
            execute_(db_file, insert_sql, tuple(mtx[i]))
        else:
            execute_(db_file, update_sql, tuple(mtx[i][-4:] + [a[0][0]]))


def update_badproto(db_file: str, mtx: List[List[str]]):
    find_sql = 'select id from t_bad_proto where those=? and problem_proto=? and not_rule=?'
    update_sql = 'update t_bad_proto set is_noti=?,note=?,prev_noti_date=?,is_need_noti=? where id=?'
    insert_sql = 'insert into t_bad_proto' \
                 '(term,those,problem_proto,not_rule,is_noti,note,prev_noti_date,is_need_noti)' \
                 ' values(?,?,?,?,?,?,?,?)'
    for i in range(1, len(mtx)):
        mtx[i][-4] = 1 if mtx[i][-4] == '是' else 0  # is_noti
        mtx[i][-1] = 0 if mtx[i][-1] == '否' else 1  # is_need_noti
        a = select_(db_file, find_sql, tuple(mtx[i][1:4]))
        if len(a) <= 0:
            execute_(db_file, insert_sql, tuple(mtx[i]))
        else:
            execute_(db_file, update_sql, (mtx[i][-4], mtx[i][-3], mtx[i][-2], mtx[i][-1], a[0][0]))


def update_storawgit(db_file: str, mtx: List[List]):
    find_sql = 'select id,is_need_noti,prev_noti_date from t_stora_wgit where module_name=? and module_those=?'
    update_sql = 'update t_stora_wgit set is_noti=?,note=?,prev_noti_date=?,is_need_noti=? where id=?'
    insert_sql = 'insert into t_stora_wgit' \
                 '(module_name,is_stora_wgit,inline_datetime,module_those,module_group,module_ori,is_noti,note,prev_noti_date,is_need_noti)' \
                 ' values(?,?,?,?,?,?,?,?,?,?)'
    for i in range(1, len(mtx)):
        mtx[i][-4] = 1 if mtx[i][-4] == '是' else 0
        mtx[i][-1] = 0 if mtx[i][-1] == '否' else 1
        a = select_(db_file, find_sql, (mtx[i][0], mtx[i][3]))
        if len(a) <= 0:
            execute_(db_file, insert_sql, tuple(mtx[i]))
        else:
            params = tuple(mtx[i][-4:] + [a[0][0]])
            execute_(db_file, update_sql, params)


def update_comacc(db_file: str, mtx: List[List]) -> List[List]:
    find_sql = 'select id,is_need_noti,prev_noti_date from t_com_acc where account=?'
    update_sql = 'update t_com_acc set is_noti=?,note=?,prev_noti_date=?,is_need_noti=? where id=?'
    insert_sql = 'insert into t_com_acc' \
                 '(account,patchbuild_times,acc_times,is_noti,note,is_need_noti,prev_noti_date)' \
                 ' values(?,?,?,?,?,?,?)'
    for i in range(1, len(mtx)):
        row = mtx[i]
        patchbuild_times = int(row[1])
        acc_times = int(row[2])
        is_noti = 1 if mtx[i][-4] == '是' else 0
        is_need_noti = 0 if mtx[i][-1] == '否' else 1
        tp = (row[0], patchbuild_times, acc_times, is_noti, row[5], is_need_noti, row[6])
        a = select_(db_file, find_sql, (mtx[i][0],))
        if len(a) <= 0:
            execute_(db_file, insert_sql, tp)
        else:
            execute_(db_file, update_sql, (is_noti, mtx[i][-3], mtx[i][-2], is_need_noti, a[0][0]))


def update_toomorede(db_file: str, mtx: List[List]) -> List[List]:
    find_sql = 'select id,is_need_noti,prev_noti_date from t_too_more_target where file=? and target_name=? and those=?'
    update_sql = 'update t_too_more_target set is_note=?,noti=?,prev_noti_date=?,is_need_noti=? where id=?'
    insert_sql = 'insert into t_too_more_target' \
                 '(file,target_name,target_count,those,is_noti,note,prev_noti_date,is_need_noti)' \
                 ' values(?,?,?,?,?,?,?,?)'
    for i in range(1, len(mtx)):
        mtx[i][-4] = 1 if mtx[i][-4] == '是' else 0
        mtx[i][-1] = 0 if mtx[i][-1] == '否' else 1
        a = select_(db_file, find_sql, (mtx[i][0], mtx[i][1], mtx[i][3]))
        if len(a) <= 0:
            execute_(db_file, insert_sql, tuple(mtx[i]))
        else:
            execute_(db_file, update_sql, tuple(mtx[i][-4:] + [a[0][0]]))


def update(db_file: str = None, in_file: str = None):
    if db_file is None:
        db_file = f'{sys.path[0]}{os.path.sep}IdealHome.db3'
    if in_file is None:
        in_date = datetime.datetime.strftime(datetime.datetime.now(), '%Y-%m-%d')
        in_file = f'{in_date}-out{os.path.sep}__out.xlsx'
    print(db_file, in_file)
    all_ = read_db_xlsx(in_file)
    update_badlib(db_file, all_.bad_lib_s)
    update_badproto(db_file, all_.bad_proto_s)
    update_comacc(db_file, all_.com_acc_s)
    update_toomorede(db_file, all_.more_tag_s)
    update_storawgit(db_file, all_.stora_wgit_s)


help_s = '''
example:
AutoWork0 genreport -d {default:IdealHome.db3} -o {default:date-out} -i {default:today}
AutoWork0 update -d {default:IdealHome.db3} -i {default:date-out\__out.xlsx}

'''


def main(args):
    db_file = None
    out_dir = None
    in_param = None
    print(args)
    if len(args) <= 0:
        print(help_s)
        exit(1)
    idx = 1
    while idx < len(args):
        if args[idx] == '-d':
            idx += 1
            db_file = args[idx]
        elif args[idx] == '-o':
            idx += 1
            out_dir = args[idx]
        elif args[idx] == '-i':
            idx += 1
            in_param = args[idx]
        else:
            print(help_s)
            exit(1)
        idx += 1

    if strip(args[0]) == 'genreport':
        genreport(db_file=db_file, out_dir=out_dir, in_date=in_param)
    elif strip(args[0]) == 'update':
        update(db_file=db_file, in_file=in_param)
    else:
        print(help_s)
        exit(1)
    exit(0)


if __name__ == '__main__':
    # main(['genreport'])
    main(sys.argv[1:])
    # genreport()
    pass
