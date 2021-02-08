from typing import List, Tuple

from openpyxl import load_workbook
from openpyxl.worksheet import worksheet

import dh_entities as dh
import sqlite3 as s3
import contextlib


@contextlib.contextmanager
def get_cursor(db: str):
    conn = None
    try:
        conn = s3.connect(db)
        cur = conn.cursor()
        yield cur
        conn.commit()
    except Exception as e:
        print(str(e))
        conn.rollback()
        raise cur
    finally:
        if conn is not None:
            conn.close()


def strip(s: str) -> str:
    return '' if s is None else str(s).strip()


def _stand_strip(s: str) -> str:
    s = strip(s)
    return s.replace('\r\n', chr(10))


def _stand_strip_br(s: str) -> str:
    s = strip(s)
    return s.replace('\n', '<tadus_br>')


def _imstand_strip_br(s: str) -> str:
    s = strip(s)
    return s.replace('<tadus_br>', '\n')


def _stand_strip_w(s: str) -> str:
    s = _stand_strip(s)
    return s.replace('\n', '\r\n')


def Yes2int(bb: str) -> int:
    x = strip(bb)
    if x == '是':
        return 1
    elif x.upper() in ['YES', 'TRUE']:
        return 1
    return 0


def No2int(bb: str) -> int:
    x = strip(bb)
    if x == '否':
        return 0
    elif x.upper() in ['NO', 'FALSE']:
        return 0
    return 1


db_file = r'C:\Users\v_fantnzeng\Desktop\notebook\理想家园-整洁GIT\每日-项目全局简报\IdealHome.db3'

db_xlsx = {
    'bad_lib': r'C:\Users\v_fantnzeng\Desktop\notebook\理想家园-整洁GIT\每日-项目全局简报\违规文件\xl_db.xlsx',
    'bad_proto': r'C:\Users\v_fantnzeng\Desktop\notebook\理想家园-整洁GIT\每日-项目全局简报\proto文件规范化检查\dest.xlsx',
    'com_acc': r'C:\Users\v_fantnzeng\Desktop\notebook\理想家园-整洁GIT\每日-项目全局简报\编译加速\need_to_noti.xlsx',
    'stora_wgit': r'C:\Users\v_fantnzeng\Desktop\notebook\理想家园-整洁GIT\每日-项目全局简报\过去10天新增模块接入整洁GIT情况\xl_db.xlsx',
    'too_more_tag': r'C:\Users\v_fantnzeng\Desktop\notebook\理想家园-整洁GIT\每日-项目全局简报\模块编译依赖明细\xl_db.xlsx'
}


def load_sht(file_path: str, sht_idx: int = 0) -> worksheet:
    wb = load_workbook(file_path)
    return wb[wb.sheetnames[sht_idx]]


def count_sht_rownum(sht: worksheet, col_num: int, begin_p: int = 2) -> int:
    count = 0
    while True:
        for i in range(col_num):
            value = strip(sht.cell(row=begin_p + count, column=i + 1).value)
            if value != '':
                count += 1
                break
        else:
            return count
    return count


def get_mtx_from_xlsx(file_path: str, col_num: int, sht_idx: int = 0) -> List[List]:
    sht = load_sht(file_path, sht_idx)
    begin_p = 2
    row_num = count_sht_rownum(sht, col_num, begin_p)
    res = list()
    for row in range(begin_p, begin_p + row_num):
        g = (_stand_strip_br(sht.cell(row=row, column=i + 1).value) for i in range(col_num))
        res.append(list(g))
    return res


def insert_many_badlib(rows: List[List]):
    def one(row: List) -> Tuple:
        values = (row[0], row[1], row[2], row[3], row[4], No2int(row[5]), row[6], No2int(row[8]), row[7])
        return values

    values_s = [one(row) for row in rows]
    sql = r'insert into t_bad_lib ' \
          r'(file,lines,not_xlib,tobe_xlib,those,is_noti,noti,is_need_noti,prev_noti_date)' \
          r' values(?,?,?,?,?,?,?,?,?);'
    with get_cursor(db_file) as cur:
        cur.executemany(sql, values_s)


def migrate_badLib():
    mtx = get_mtx_from_xlsx(db_xlsx['bad_lib'], 9)
    insert_many_badlib(mtx)


def migrate_badproto():
    mtx = get_mtx_from_xlsx(db_xlsx['bad_proto'], 8)

    def one(row: List) -> Tuple:
        return (row[0], row[1], row[2], row[3], No2int(row[4]), row[5], No2int(row[6]), row[7])

    values_s = [one(row) for row in mtx]
    sql = 'insert into t_bad_proto' \
          '(term,those,problem_proto,not_rule,is_noti,note,is_need_noti,prev_noti_date)' \
          ' values(?,?,?,?,?,?,?,?)'
    with get_cursor(db_file) as cur:
        cur.executemany(sql, values_s)


def migrate_com_acc():
    mtx = get_mtx_from_xlsx(db_xlsx['com_acc'], 8)

    def one(row: List) -> Tuple:
        try:
            patchbuild_times = int(row[1])
            acc_times = int(row[2])
        except Exception as e:
            patchbuild_times = 0
            acc_times = 0
        return (row[0], patchbuild_times, acc_times, No2int(row[4]), row[5], No2int(row[7]), row[6])

    values_s = [one(row) for row in mtx]
    sql = 'insert into t_com_acc' \
          '(account,patchbuild_times,acc_times,is_noti,note,is_need_noti,prev_noti_date)' \
          ' values(?,?,?,?,?,?,?)'

    with get_cursor(db_file) as cur:
        cur.executemany(sql, values_s)


def migrate_stora_wgit():
    mtx = get_mtx_from_xlsx(db_xlsx['stora_wgit'], 10)

    def one(row: List) -> Tuple:
        is_in_wgit = 1 if row[1] == '已入整洁GIT' else 0
        return (row[0], is_in_wgit, row[2], row[3], row[4], row[5], No2int(row[6]), row[7], No2int(row[9]), row[8])

    values_s = [one(row) for row in mtx]
    sql = 'insert into t_stora_wgit' \
          '(module_name,is_stora_wgit,inline_datetime,module_those,module_group,module_ori,is_noti,note,is_need_noti,prev_noti_date)' \
          ' values(?,?,?,?,?,?,?,?,?,?)'

    with get_cursor(db_file) as cur:
        cur.executemany(sql, values_s)


def migrate_too_more_de():
    mtx = get_mtx_from_xlsx(db_xlsx['too_more_tag'], 8)

    def one(row: List) -> Tuple:
        try:
            target_count = int(row[2])
        except Exception as e:
            target_count = 0

        return (row[0], row[1], target_count, row[3], No2int(row[4]), row[5], No2int(row[7]), row[6])

    values_s = [one(row) for row in mtx]
    sql = 'insert into t_too_more_target' \
          '(file,target_name,target_count,those,is_noti,note,is_need_noti,prev_noti_date)' \
          ' values(?,?,?,?,?,?,?,?)'

    with get_cursor(db_file) as cur:
        cur.executemany(sql, values_s)


if __name__ == '__main__':
    migrate_badLib()
    migrate_badproto()
    migrate_com_acc()
    migrate_stora_wgit()
    migrate_too_more_de()
    pass
